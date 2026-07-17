"""Bounded CSV/XLSX utilities with formula-injection and archive defenses."""

from __future__ import annotations

import csv
import io
import re
import zipfile
from collections.abc import Iterator, Sequence
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path, PurePosixPath
from typing import Any
from xml.etree.ElementTree import ParseError

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.schemas.employee_import import EMPLOYEE_IMPORT_FIELDS, EMPLOYEE_IMPORT_MAX_ROWS

_FORMULA_PREFIXES = frozenset({"=", "+", "-", "@", "\t", "\r", "\n"})
_ILLEGAL_SPREADSHEET_CONTROLS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
_XLSX_MAX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
_XLSX_MAX_ENTRY_BYTES = 20 * 1024 * 1024
_XLSX_MAX_ENTRIES = 1_000


class SpreadsheetFileError(ValueError):
    def __init__(self, code: str) -> None:
        super().__init__(code)
        self.code = code


def spreadsheet_safe(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        rendered = value.isoformat()
    elif isinstance(value, date):
        rendered = value.isoformat()
    elif isinstance(value, Decimal):
        rendered = format(value, "f")
    else:
        rendered = str(value)
    rendered = _ILLEGAL_SPREADSHEET_CONTROLS.sub("\ufffd", rendered)
    candidate = rendered.lstrip()
    if candidate and candidate[0] in _FORMULA_PREFIXES:
        return f"'{rendered}"
    return rendered


def employee_import_template_csv() -> bytes:
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, lineterminator="\r\n")
    writer.writerow(EMPLOYEE_IMPORT_FIELDS)
    return ("\ufeff" + buffer.getvalue()).encode("utf-8")


def employee_import_template_xlsx() -> bytes:
    workbook = Workbook(write_only=False)
    sheet = workbook.active
    sheet.title = "employees"
    sheet.append(list(EMPLOYEE_IMPORT_FIELDS))
    sheet.freeze_panes = "A2"
    instructions = workbook.create_sheet("instructions")
    instructions.append(["template_version", "1"])
    instructions.append(["maximum_rows", "10000"])
    instructions.append(["date_format", "YYYY-MM-DD"])
    instructions.append(["allowed_statuses", "active,on_leave"])
    instructions.append(["employment_end_date", "blank_in_v1"])
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def iter_import_rows(path: Path, file_format: str) -> Iterator[tuple[int, dict[str, str]]]:
    if file_format == "csv":
        yield from _iter_csv_rows(path)
        return
    if file_format == "xlsx":
        yield from _iter_xlsx_rows(path)
        return
    raise SpreadsheetFileError("invalid_file")


def _iter_csv_rows(path: Path) -> Iterator[tuple[int, dict[str, str]]]:
    previous_limit = csv.field_size_limit()
    csv.field_size_limit(1_000_000)
    try:
        with path.open("r", encoding="utf-8-sig", errors="strict", newline="") as handle:
            reader = csv.reader(handle, strict=True)
            try:
                header = next(reader)
            except StopIteration as exc:
                raise SpreadsheetFileError("invalid_headers") from exc
            _validate_header(header)
            for row_number, cells in enumerate(reader, start=2):
                if row_number > EMPLOYEE_IMPORT_MAX_ROWS + 1:
                    raise SpreadsheetFileError("row_limit_exceeded")
                if len(cells) != len(EMPLOYEE_IMPORT_FIELDS):
                    raise SpreadsheetFileError("invalid_row_shape")
                if not any(cell.strip() for cell in cells):
                    continue
                if any("\x00" in cell for cell in cells):
                    raise SpreadsheetFileError("invalid_file")
                yield row_number, dict(zip(EMPLOYEE_IMPORT_FIELDS, cells, strict=True))
    except (UnicodeDecodeError, csv.Error) as exc:
        raise SpreadsheetFileError("invalid_file") from exc
    finally:
        csv.field_size_limit(previous_limit)


def _iter_xlsx_rows(path: Path) -> Iterator[tuple[int, dict[str, str]]]:
    _validate_xlsx_archive(path)
    try:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except (
        InvalidFileException,
        KeyError,
        OSError,
        ParseError,
        TypeError,
        ValueError,
        zipfile.BadZipFile,
    ) as exc:
        raise SpreadsheetFileError("invalid_file") from exc
    try:
        if "employees" not in workbook.sheetnames:
            raise SpreadsheetFileError("invalid_headers")
        sheet = workbook["employees"]
        if (sheet.max_row or 0) > EMPLOYEE_IMPORT_MAX_ROWS + 1:
            raise SpreadsheetFileError("row_limit_exceeded")
        if (sheet.max_column or 0) > len(EMPLOYEE_IMPORT_FIELDS):
            raise SpreadsheetFileError("invalid_row_shape")
        rows = sheet.iter_rows(
            min_col=1,
            max_col=len(EMPLOYEE_IMPORT_FIELDS),
            values_only=True,
        )
        try:
            header = [_cell_text(value) for value in next(rows)]
        except StopIteration as exc:
            raise SpreadsheetFileError("invalid_headers") from exc
        _validate_header(header)
        for row_number, values in enumerate(rows, start=2):
            cells = [_cell_text(value) for value in values]
            if len(cells) > len(EMPLOYEE_IMPORT_FIELDS):
                if any(cell.strip() for cell in cells[len(EMPLOYEE_IMPORT_FIELDS) :]):
                    raise SpreadsheetFileError("invalid_row_shape")
                cells = cells[: len(EMPLOYEE_IMPORT_FIELDS)]
            cells.extend([""] * (len(EMPLOYEE_IMPORT_FIELDS) - len(cells)))
            if not any(cell.strip() for cell in cells):
                continue
            yield row_number, dict(zip(EMPLOYEE_IMPORT_FIELDS, cells, strict=True))
    except SpreadsheetFileError:
        raise
    except (
        InvalidFileException,
        KeyError,
        OSError,
        ParseError,
        TypeError,
        ValueError,
        zipfile.BadZipFile,
    ) as exc:
        raise SpreadsheetFileError("invalid_file") from exc
    finally:
        workbook.close()


def _validate_header(header: Sequence[str]) -> None:
    normalized = [value.strip() for value in header]
    if normalized != list(EMPLOYEE_IMPORT_FIELDS):
        raise SpreadsheetFileError("invalid_headers")


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return (
            value.date().isoformat()
            if value.time().isoformat() == "00:00:00"
            else value.isoformat()
        )
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def _validate_xlsx_archive(path: Path) -> None:
    try:
        with zipfile.ZipFile(path) as archive:
            entries = archive.infolist()
            if len(entries) > _XLSX_MAX_ENTRIES:
                raise SpreadsheetFileError("invalid_file")
            total = 0
            for entry in entries:
                pure_path = PurePosixPath(entry.filename)
                if pure_path.is_absolute() or ".." in pure_path.parts or "\\" in entry.filename:
                    raise SpreadsheetFileError("invalid_file")
                if entry.file_size > _XLSX_MAX_ENTRY_BYTES:
                    raise SpreadsheetFileError("invalid_file")
                total += entry.file_size
                if total > _XLSX_MAX_UNCOMPRESSED_BYTES:
                    raise SpreadsheetFileError("invalid_file")
    except zipfile.BadZipFile as exc:
        raise SpreadsheetFileError("invalid_file") from exc


__all__ = [
    "SpreadsheetFileError",
    "employee_import_template_csv",
    "employee_import_template_xlsx",
    "iter_import_rows",
    "spreadsheet_safe",
]
